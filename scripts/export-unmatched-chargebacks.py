import csv
import json
import re
import unicodedata
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path("/Users/vMac/01_projects/Email Marketing/resend-takeover/data/current")
CONTACTS = json.loads((ROOT / "hubspot-snapshot/raw/contacts.json").read_text())
LEDGER = json.loads((ROOT / "ledger/canonical-ledger.json").read_text())
OUTPUT = ROOT / "review/unmatched-chargeback-dispute-payments.csv"

def normalize_email(value):
    return str(value or "").strip().lower()

def split_emails(value):
    return [normalize_email(item) for item in re.split(r"[;,\s]+", str(value or "")) if "@" in item]

def normalize_name(value):
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode().lower()
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", text)).strip()

def name_tokens(value):
    return tuple(sorted(normalize_name(value).split()))

current_emails = defaultdict(list)
current_names = defaultdict(list)
current_name_strings = []
for contact in CONTACTS:
    properties = contact.get("properties", {})
    for field in ("email", "hs_additional_emails", "legacy_email", "work_email", "hs_conversations_visitor_email"):
        for address in split_emails(properties.get(field)):
            current_emails[address].append(contact)
    names = []
    for first, last in ((properties.get("firstname"), properties.get("lastname")), (properties.get("legacy_firstname"), properties.get("legacy_lastname"))):
        full_name = normalize_name(f"{first or ''} {last or ''}")
        if len(full_name) >= 4:
            names.append(full_name)
            current_names[name_tokens(full_name)].append(contact)
    current_name_strings.append((contact, names))

ledger_emails = defaultdict(list)
ledger_names = defaultdict(list)
for record in LEDGER:
    if normalize_email(record.get("email")):
        ledger_emails[normalize_email(record["email"])].append(record)
    full_name = normalize_name(f"{record.get('first_name') or ''} {record.get('last_name') or ''}")
    if full_name:
        ledger_names[name_tokens(full_name)].append(record)

def current_match(email, name):
    email_matches = current_emails.get(normalize_email(email), [])
    if len(email_matches) == 1:
        return True
    exact_name_matches = current_names.get(name_tokens(name), [])
    if len(exact_name_matches) == 1:
        return True
    normalized = normalize_name(name)
    if not normalized:
        return False
    scores = []
    for contact, aliases in current_name_strings:
        score = max((SequenceMatcher(None, normalized, alias).ratio() for alias in aliases), default=0)
        if score >= 0.82:
            scores.append(score)
    scores.sort(reverse=True)
    return bool(scores and scores[0] >= 0.90 and (len(scores) == 1 or scores[0] - scores[1] >= 0.08))

def resolve_current(email, name):
    email_matches = current_emails.get(normalize_email(email), [])
    if len(email_matches) == 1:
        return email_matches[0], "exact_or_alias_email"
    exact_name_matches = current_names.get(name_tokens(name), [])
    if len(exact_name_matches) == 1:
        return exact_name_matches[0], "exact_normalized_name"
    normalized = normalize_name(name)
    if not normalized:
        return None, None
    scores = []
    for contact, aliases in current_name_strings:
        score = max((SequenceMatcher(None, normalized, alias).ratio() for alias in aliases), default=0)
        if score >= 0.82:
            scores.append((score, contact))
    scores.sort(key=lambda item: item[0], reverse=True)
    if scores and scores[0][0] >= 0.90 and (len(scores) == 1 or scores[0][0] - scores[1][0] >= 0.08):
        return scores[0][1], "high_confidence_fuzzy_name"
    return None, None

def historic_match(email, name):
    return len(ledger_emails.get(normalize_email(email), [])) == 1 or len(ledger_names.get(name_tokens(name), [])) == 1

source_rows = []
workbook = load_workbook("/Users/vMac/Downloads/chargebacks.xlsx", read_only=True, data_only=True)
sheet = workbook["chargebacks"]
headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
for values in sheet.iter_rows(min_row=2, values_only=True):
    row = dict(zip(headers, values))
    source_rows.append({
        "source_document": "chargebacks.xlsx",
        "source_customer_name": row.get("customer_name", ""),
        "source_customer_email": row.get("email", ""),
        "source_row": row
    })
with open("/Users/vMac/Downloads/unified_payments.csv", newline="", encoding="utf-8-sig") as handle:
    for row in csv.DictReader(handle):
        disputed_amount = str(row.get("Disputed Amount") or "").strip()
        dispute_date = str(row.get("Dispute Date (UTC)") or "").strip()
        dispute_status = str(row.get("Dispute Status") or "").strip()
        if disputed_amount not in ("", "0", "0.0", "0.00") or dispute_date or dispute_status:
            source_rows.append({
                "source_document": "unified_payments.csv",
                "source_customer_name": row.get("Shipping Name") or row.get("Card Name", ""),
                "source_customer_email": row.get("Customer Email", ""),
                "source_row": row
            })

chargeback_ids = {str(row["source_row"].get("charge_id_or_payment_id") or "") for row in source_rows}
enrichment_specs = [
    ("stripe_transactions_itemized", "/Users/vMac/Downloads/Stripe Transactions.xlsx", "itemized_balance_change_from_ac", ("charge_id",)),
    ("stripe_charges_master", "/Users/vMac/Downloads/Stripe Charges Masterfile.xlsx", "Stripe Charges 22af4e0d84e08025", ("Charge ID",)),
    ("stripe_charges_payout_reconciliation", "/Users/vMac/Downloads/Stripe Charges Masterfile.xlsx", "itemized_payout_reconciliation_", ("charge_id", "payment_intent_id")),
    ("stripe_charges_balance_detail", "/Users/vMac/Downloads/Stripe Charges Masterfile.xlsx", "Sheet2", ("charge_id", "payment_intent_id"))
]
enrichment_maps = {}
enrichment_columns = []
for prefix, workbook_path, sheet_name, key_columns in enrichment_specs:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    columns = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    enrichment_columns.extend([f"{prefix}__{column}" for column in columns if column])
    matched_rows = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        raw = dict(zip(columns, values))
        for key_column in key_columns:
            key = str(raw.get(key_column) or "")
            if key in chargeback_ids and key not in matched_rows:
                matched_rows[key] = raw
    enrichment_maps[prefix] = (columns, matched_rows)
all_source_columns = []
for row in source_rows:
    for column in row["source_row"]:
        if column not in all_source_columns:
            all_source_columns.append(column)
OUTPUT.parent.mkdir(parents=True, exist_ok=True)
MATCHED_OUTPUT = ROOT / "review/chargeback-dispute-payments-matched-to-hubspot.csv"
PROPOSED_UPDATES_OUTPUT = ROOT / "review/chargeback-customer-situation-proposed-updates.csv"
headers = ["source_document", "source_customer_name", "source_customer_email", "hubspot_match_result", "hubspot_match_basis", "hubspot_contact_id", "hubspot_primary_email", "hubspot_customer_situation", *all_source_columns, *enrichment_columns]
matched_rows = []
unmatched_rows = []
for row in source_rows:
    output = {
            "source_document": row["source_document"],
            "source_customer_name": row["source_customer_name"],
            "source_customer_email": row["source_customer_email"],
            **row["source_row"]
    }
    chargeback_id = str(row["source_row"].get("charge_id_or_payment_id") or "")
    for prefix, (columns, matches) in enrichment_maps.items():
        source = matches.get(chargeback_id, {})
        for column in columns:
            if column:
                output[f"{prefix}__{column}"] = source.get(column, "")
    enriched_email = output.get("stripe_transactions_itemized__customer_email") or output.get("stripe_charges_master__Customer Email") or output.get("stripe_charges_balance_detail__customer_email") or row["source_customer_email"]
    enriched_name = output.get("stripe_transactions_itemized__customer_name") or output.get("stripe_charges_master__Customer Name") or output.get("stripe_charges_balance_detail__customer_name") or row["source_customer_name"]
    contact, basis = resolve_current(enriched_email, enriched_name)
    if contact:
        properties = contact.get("properties", {})
        output.update({
            "hubspot_match_result": "matched_current_hubspot_contact",
            "hubspot_match_basis": basis,
            "hubspot_contact_id": contact.get("id", ""),
            "hubspot_primary_email": properties.get("email", ""),
            "hubspot_customer_situation": properties.get("customer_situation", "")
        })
        matched_rows.append(output)
    else:
        output["hubspot_match_result"] = "no match in current portal or historic exports after email, aliases, normalized-name, and fuzzy-name checks"
        unmatched_rows.append(output)
for target, rows in ((OUTPUT, unmatched_rows), (MATCHED_OUTPUT, matched_rows)):
    with target.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)
proposed_updates = {}
for row in matched_rows:
    current = str(row.get("hubspot_customer_situation") or "")
    current_options = [option for option in current.split(";") if option]
    if "Chargeback" in current_options:
        continue
    proposed_updates[row["hubspot_contact_id"]] = {
        "hubspot_contact_id": row["hubspot_contact_id"],
        "hubspot_primary_email": row["hubspot_primary_email"],
        "current_customer_situation": current,
        "proposed_customer_situation": ";".join([*current_options, "Chargeback"]),
        "payment_rows_confirming_dispute": sum(1 for matched in matched_rows if matched["hubspot_contact_id"] == row["hubspot_contact_id"]),
        "match_basis": row["hubspot_match_basis"]
    }
update_headers = ["hubspot_contact_id", "hubspot_primary_email", "current_customer_situation", "proposed_customer_situation", "payment_rows_confirming_dispute", "match_basis"]
with PROPOSED_UPDATES_OUTPUT.open("w", newline="", encoding="utf-8") as handle:
    writer = csv.DictWriter(handle, fieldnames=update_headers)
    writer.writeheader()
    writer.writerows(sorted(proposed_updates.values(), key=lambda row: row["hubspot_primary_email"].lower()))
print(json.dumps({"unmatched_output": str(OUTPUT), "matched_output": str(MATCHED_OUTPUT), "proposed_updates_output": str(PROPOSED_UPDATES_OUTPUT), "matched_payment_rows": len(matched_rows), "unmatched_payment_rows": len(unmatched_rows), "proposed_customer_situation_updates": len(proposed_updates)}))
